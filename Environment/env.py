import pygame          # 导入模块
from pygame.locals import *    # 导入pygame一些常用的函数和常量
import pandas as pd
# openpyxl index starts from 1
from openpyxl import load_workbook
from Controller import solution_random
from Model import agent
from Common import parameter as para
from Map import excel_trans
import random
import copy
import operator
pygame.init()
info = pygame.font.SysFont('rachana', 25)


class Env_Base:
    def __init__(self, bus_number, package_car_number, advance_time):
        # attributes about environment map information
        self.path_file = open(para.path_file, "r")
        self.map_info = excel_trans.Map_info(eval(self.path_file.readline()))
        self.task_data = load_workbook(para.task_data)
        self.task_sheets = self.task_data.sheetnames
        # TODO: from  0 to i to train all sheets
        self.task = self.task_data[self.task_sheets[0]]

        # uncompleted task group
        self.wait_queue = []

        self.bus_group = pygame.sprite.Group()
        self.pc_group = pygame.sprite.Group()
        self.bus_avai_group = pygame.sprite.Group()
        self.pc_avai_group = pygame.sprite.Group()

        self.bus_n = bus_number
        self.pc_n = package_car_number
        # the advanced time to make dispatching decision
        self.advance_t = para.advance_time
        # the time of starting simulation
        self.time = self.task.cell(2, 2) - self.advance_t
        self.first_index = 2


    def init_env(self):
        self.create_bus_pc()

    # create the map and path for the environment
    def create_map(self):
        pass

    # create bus and pc for the environment
    def create_bus_pc(self):
        for i in range(self.bus_n):
            if i < int(len(self.bus_n) / 2):
                locals()['bus' + str(i)] = agent.Agent_Base(self.list_bus_p[0], 0)
            else:
                locals()['bus' + str(i)] = agent.Agent_Base(self.list_bus_p[1], 0)
            self.bus_group.add(locals()['bus' + str(i)])
            self.bus_avai_group.add(locals()['bus' + str(i)])

        for j in range(self.pc_n):
            if j < int(len(self.bus_n) / 2):
                locals()['pc' + str(j)] = agent.Agent_Base(self.list_bus_p[2], 1)
            else:
                locals()['pc' + str(j)] = agent.Agent_Base(self.list_bus_p[3], 1)
            self.pc_group.add(locals()['pc' + str(j)])
            self.pc_avai_group.add(locals()['pc' + str(j)])

    def step(self):
        for i in range(self.first_index, self.task.max_row+1):
            if self.task.cell(i,2).value-self.time == self.advance_t:
                self.wait_queue.append(i)
            else:
                self.first_index = i
                break
        # TODO: consider the condition when there is no available vehicle to do tasks
        for item in self.wait_queue:
            if self.task.cell(item, 7).value == 'shuttle':
                vehicle = solution_random.schedule_bus(self.bus_avai_group)
            else:
                vehicle = solution_random.schedule_pc(self.pc_avai_group)
            path = self.map_info.get_map_info(vehicle.get_place, eval(self.task.cell(item,5).value), eval(self.task.cell(item,6).value))
            vehicle.get_schedule(path)
            # TODO:write a controller to decide the time of vehicle departure
            vehicle.t_start = self.time + None
            self.wait_queue.remove(item)

    def get_observation(self):
        # TODO: wait for writing
        pass



class Env(Env_Base):
    def __init__(self, bus_number, package_car_number, advance_time):
        super().__init__(bus_number, package_car_number, advance_time)
        # initial the screen & background image
        self.screen = pygame.display.set_mode((para.width, para.height))
        self.bg = pygame.image.load(para.bg_name)
        # put bg on the assigned point in the memory,the bg is not showed in this step
        self.screen.blit(self.bg, (0, 0))  # (x, y) is the left up location
        self.mark_group = pygame.sprite.Group()


    def init_env(self):

        self.create_map()
        # when the simulation needs visualization, the following two functions will be rewritten.
        self.create_bus_pc()
        self.create_landmark()

    # create bus and pc for the environment
    def create_bus_pc(self):
        for i in range(self.bus_n):
            if i < int(len(self.bus_n) / 2):
                locals()['bus' + str(i)] = agent.Agent(para.bus_img, self.list_bus_p[0], 0)
            else:
                locals()['bus' + str(i)] = agent.Agent(para.bus_img, self.list_bus_p[1], 0)
            self.bus_group.add(locals()['bus' + str(i)])
            self.bus_avai_group.add(locals()['bus' + str(i)])

        for j in range(self.pc_n):
            if j < int(len(self.bus_n) / 2):
                locals()['pc' + str(j)] = agent.Agent(para.pc_img, self.list_bus_p[2], 1)
            else:
                locals()['pc' + str(j)] = agent.Agent(para.pc_img, self.list_bus_p[3], 1)
            self.pc_group.add(locals()['pc' + str(j)])
            self.pc_avai_group.add(locals()['pc' + str(j)])

        # put the land_mark on the screen to display
    def create_landmark(self):
        j = 0
        for j in range(len(self.list_near_p)):
            locals()['mark' + str(j)] = agent.Agent(para.passenger_near_img, self.list_near_p[j], 2)
            self.mark_group.add(locals()['mark' + str(j)])
        number = j
        for j in range(number + 1, number + len(self.list_far_p) + 1):
            locals()['mark' + str(j)] = agent.Agent(para.park_point_img, self.list_far_p[j - number - 1], 2)
            self.mark_group.add(locals()['mark' + str(j)])
        locals()['mark' + str(j + 1)] = agent.Agent(para.package_center_img, self.package_center, 2)
        self.mark_group.add(locals()['mark' + str(j + 1)])

    # update the visualization of the environment
    def env_render(self):
        '''
                :param agent_group: the shuffle and package car group
                :param mark_group: the land mark group
                :return:
                '''
        self.screen.blit(self.bg, (0, 0))  # update the background image
        self.create_landmark()
        # update the display of logos
        # self.set_landmark()

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                print("end simulation...")
                pygame.quit()
                exit(0)
        # update the display of cars and landmarks
        self.mark_group.draw(self.screen)
        self.bus_group.update(self.time)
        self.pc_group.update(self.time)
        self.bus_group.draw(self.screen)
        self.pc_group.draw(self.screen)

        # update the screen
        pygame.display.update()
        info_fmt = info.render("time:" + str(self.time), True, (0, 0, 0))
        self.screen.blit(info_fmt, (0, 0))
        # 更新屏幕
        pygame.display.flip()
        pygame.time.delay(100)
        # environment render once, env time +1