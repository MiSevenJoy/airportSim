import random
import sys

import gym
import pygame
from gym.envs.classic_control import rendering
# openpyxl index starts from 1
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter, column_index_from_string
from Model import agent
from Common import parameter as para
from Map import excel_trans
from Controller import solution_random as sr


class Env_Base(gym.Env):
    def __init__(self):
        # attributes about environment map information
        self.path_file = open(para.path_file, "r")
        self.path_list = self.path_file.readlines()
        self.vehicle_parks = eval(self.path_list[3])
        self.map_info = excel_trans.Map_info(self.path_list)
        self.task_data = load_workbook(para.task_data)
        self.task_sheets = self.task_data.sheetnames
        self.task = self.task_data[self.task_sheets[0]]
        # self.task = None

        # uncompleted task group
        self.wait_queue = []
        # vehicle group
        self.bus_avai_group = pygame.sprite.Group()
        self.pc_avai_group = pygame.sprite.Group()
        self.bus_busy_group = pygame.sprite.Group()
        self.pc_busy_group = pygame.sprite.Group()

        self.bus_n = para.n_shuttle
        self.pc_n = para.n_luggage_van
        # the advanced time to make dispatching decision
        self.advance_t = para.advance_time
        # the time of starting simulation
        self.time = self.task.cell(2, 3).value - self.advance_t
        self.first_index = 2
        self.dispatched_agv = 0

    def reset(self):
        self.__init__()
        self.create_bus_pc()
        print('complete environment reset.')

    # create the map and path for the environment
    def create_map(self):
        pass

    # create bus and pc for the environment
    def create_bus_pc(self):
        for i in range(self.bus_n):
            if i < int(self.bus_n / 2):
                locals()['bus' + str(i)] = agent.Agent_Base('bus' + str(i), self.vehicle_parks[0])
            else:
                locals()['bus' + str(i)] = agent.Agent_Base('bus' + str(i), self.vehicle_parks[1])
            self.bus_avai_group.add(locals()['bus' + str(i)])

        for j in range(self.pc_n):
            if j < int(self.bus_n / 2):
                locals()['pc' + str(j)] = agent.Agent_Base('pc' + str(j), self.vehicle_parks[2])
            else:
                locals()['pc' + str(j)] = agent.Agent_Base('pc' + str(j), self.vehicle_parks[3])
            self.pc_avai_group.add(locals()['pc' + str(j)])

    def env_react(self, item,
                  avai_group: pygame.sprite.Group,
                  busy_group: pygame.sprite.Group, vehicle: agent.Agent_Base):
        # vehicle = solution_random.schedule_vehicle(avai_group)
        avai_group.remove(vehicle)
        busy_group.add(vehicle)
        distance1, distance = self.map_info.get_map_info(vehicle.place, eval(self.task.cell(item, 5).value),
                                          eval(self.task.cell(item, 6).value))
        t_arrive1 = self.time + int(distance1/10)
        t_arrive2 = self.time+int(distance/10)
        # the fight departure
        if int(self.task.cell(item, 4).value)==1:
            # if vehicle arrived the aircraft lot earlier than the aircraft
            if t_arrive2 < int(self.task.cell(item, 3).value):
                total_t = int(self.task.cell(item, 3).value)-self.time
            else:
                total_t = int(distance/10)
            delay_t = t_arrive2-int(self.task.cell(item, 3).value)
        # the flight arrive
        else:
            if t_arrive1 < int(self.task.cell(item, 3).value):
                total_t = int(self.task.cell(item, 3).value)-self.time+t_arrive2-t_arrive1
            else:
                total_t = int(distance/10)
            delay_t = t_arrive1-int(self.task.cell(item, 3).value)
        vehicle.get_dispatch(self.time+total_t, eval(self.task.cell(item, 6).value))

        return delay_t, distance

    def check_vehicles_state(self, busy_vehicle_group: pygame.sprite.Group, avai_vehicle_group: pygame.sprite.Group,a,b):
        '''
        if the car has finished its task and reached the end, makes them free again
        '''
        for carAgent in busy_vehicle_group:
            # if the car just arrived the end
            if carAgent.state == 2:
                if carAgent.last_p in self.vehicle_parks:
                    carAgent.state = 0
                busy_vehicle_group.remove(carAgent)
                avai_vehicle_group.add(carAgent)
        for car in avai_vehicle_group:
            if car.state == 2 and self.time-car.t_end>1:
                p_end = self.vehicle_parks[random.randint(a,b)]
                distance = self.map_info.get_2p_info(car.place, p_end)
                # print('go back to parking lot')
                car.get_dispatch(self.time+int(distance/10), p_end)
                busy_vehicle_group.add(car)
                avai_vehicle_group.remove(car)

    def insert_listdata_to_column(self, sheet, listdata, column_name, start_row=2):
        colindex = column_index_from_string(column_name)
        for rowindex in range(start_row, start_row + len(listdata)):
            val = listdata[rowindex - start_row]
            try:
                sheet.cell(row=rowindex, column=colindex, value=val)
            except:
                val = ILLEGAL_CHARACTERS_RE.sub(r'', val)
                sheet.cell(row=rowindex, column=colindex, value=val)

    def record_all(self):
        pass
        # self.insert_listdata_to_column(self.task, self.vehicle_data, 'serve vehicle number')
        # self.insert_listdata_to_column(self.task, self.distance_data, 'distance')

    def step(self):

        for i in range(self.first_index, self.task.max_row+1):
            if self.task.cell(i, 3).value-self.time == self.advance_t:
                self.wait_queue.append(i)
                if i == self.task.max_row: self.first_index = i
            else:
                self.first_index = i
                break
        # print('len wait_queue:  '+str(len(self.wait_queue)))
        # consider the condition when there is no available vehicle to do tasks
        global item
        del_list = []
        for item in self.wait_queue:
            if len(self.bus_avai_group) > 0:
                if self.task.cell(item, 7).value == 'shuttle':
                    vehicle = sr.schedule_vehicle(self.bus_avai_group)
                    # print(vehicle)
                    reward = self.env_react(item, self.bus_avai_group, self.bus_busy_group, vehicle)
                    sr.train(reward)
                    del_list.append(item)
            else:
                print('no available shuttle.')
            if len(self.pc_avai_group) > 0:
                if self.task.cell(item, 7).value != 'shuttle':
                    vehicle = sr.schedule_vehicle(self.pc_avai_group)
                    # print(vehicle)
                    reward = self.env_react(item, self.pc_avai_group, self.pc_busy_group, vehicle)
                    sr.train(reward)
                    del_list.append(item)
            else:
                print('no available luggage van.')

            # if len(self.bus_avai_group) + len(self.pc_avai_group)==0:
            #     print('now there is not free agv')
            #     break
            # # once one task is dealt, DQN update once.
            # elif self.task.cell(item, 7).value == 'shuttle' and len(self.bus_avai_group) > 0:
            #     vehicle = sr.schedule_vehicle(self.bus_avai_group)
            #     # print(vehicle)
            #     reward = self.env_react(item, self.bus_avai_group, self.bus_busy_group, vehicle)
            #     sr.train(reward)
            #     del_list.append(item)
            # elif self.task.cell(item, 7).value != 'shuttle' and len(self.pc_avai_group) > 0:
            #     vehicle = sr.schedule_vehicle(self.pc_avai_group)
            #     reward = self.env_react(item, self.pc_avai_group, self.pc_busy_group, vehicle)
            #     sr.train(reward)
            #     del_list.append(item)

        for del_element in del_list:
            self.wait_queue.remove(del_element)

        # monitor states of vehicles which has reached the end
        self.check_vehicles_state(self.bus_busy_group, self.bus_avai_group, 0, 1)
        self.check_vehicles_state(self.pc_busy_group, self.pc_avai_group, 2, 3)

        # update the dispatched vehicles' states
        self.bus_busy_group.update(self.time)
        self.pc_busy_group.update(self.time)
        if self.time < 2000:
            print(self.time)
            print('avai agv n: ' + str(len(self.bus_avai_group) + len(self.pc_avai_group)))
        # print("now the clock is"+str(self.time))
        self.time += 1
        print('\n')

    def done(self):
        '''
        if all tasks of this day are done and all AGVs has returned to their parking lot, the simulation will end.
        '''
        if len(self.wait_queue)== 0 and self.first_index == self.task.max_row and len(self.bus_avai_group)==self.bus_n \
                and len(self.pc_avai_group)==self.pc_n:
            # print(self.time)
            print('simulation ends.')
            sys.exit(1)

